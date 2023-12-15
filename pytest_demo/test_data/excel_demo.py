import openpyxl

book = openpyxl.load_workbook("D:\\Software Testing\\python_demo.xlsx")
sheet = book.active
Dict = {}
# cell = sheet.cell(row=1, column=2)
#
# print(cell.value)
#
# sheet.cell(row=2, column=2).value = "Gabriel"  # assigning value to given cell
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)        # print the number of max columns/rows
# print(sheet.max_column)
#
# print(sheet["A5"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":    # printing specific row and column
        for j in range(2, sheet.max_column+1):

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)

