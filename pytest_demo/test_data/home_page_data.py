import openpyxl


class HomePageData:

    test_home_page_data = [{"name": "Tvoja mac", "email": "abcd@gmail.com", "gender": "Female"},
                           {"name": "Tvoj ocec", "email": "abcdefgh@gmail.com", "gender": "Male"}]

    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("D:\\Software Testing\\python_demo.xlsx")
        sheet = book.active
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # printing specific row and column
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
